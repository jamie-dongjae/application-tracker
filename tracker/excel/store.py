"""ExcelStore: the single owner of the tracker workbook.

Records are held in memory as lists of dicts; every mutation rewrites the
data sheets in full and saves atomically (tmp file + os.replace). External
edits made directly in Excel are picked up via an mtime/size signature check
before reads. Writes are refused while Excel holds the file open (detected
via the `~$` owner sentinel), surfaced to the API as HTTP 409.
"""
from __future__ import annotations

import os
import shutil
import threading
import time
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel

from . import schema

BACKUP_KEEP = 20
BACKUP_MIN_INTERVAL = 3600  # seconds


class WorkbookLockedError(Exception):
    """The workbook is open in Excel (or otherwise write-locked)."""


class NotFoundError(Exception):
    pass


def coerce_date(value) -> date | None:
    """Accept datetime/date, Excel serial numbers, and ISO-ish strings."""
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)):
        try:
            coerced = from_excel(value)
            return coerced.date() if isinstance(coerced, datetime) else coerced
        except (ValueError, TypeError, OverflowError):
            return None
    text = str(value).strip()
    for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%m/%d/%Y", "%d %b %Y"):
        try:
            return datetime.strptime(text[:10] if fmt == "%Y-%m-%d" else text, fmt).date()
        except ValueError:
            continue
    return None


def _clean(record: dict, keys: list) -> dict:
    """Normalize one in-memory record: keep known keys, coerce types."""
    out = {}
    for key in keys:
        value = record.get(key)
        if isinstance(value, str):
            value = value.strip()
        if value in (None, ""):
            out[key] = ""
            continue
        if key == "date_applied":
            out[key] = coerce_date(value) or ""
        elif key == "status":
            text = str(value).strip()
            out[key] = schema.STATUS_MIGRATE.get(text, text)
        elif key in schema.NUMERIC_KEYS:
            try:
                out[key] = float(value)
            except (TypeError, ValueError):
                out[key] = ""
        elif key == "id":
            out[key] = int(value)
        else:
            out[key] = str(value) if not isinstance(value, str) else value
    return out


_STAR_HEADERS = {"Situation": "situation", "Task": "task", "Action": "action",
                 "Result": "result", "Tips": "tips"}


def _legacy_prep_columns(ws) -> dict:
    """Map column index -> STAR key for pre-v3 prep sheets (else empty)."""
    mapping = {}
    for idx, cell in enumerate(ws[1], start=1):
        header = str(cell.value or "").strip()
        if header in _STAR_HEADERS:
            mapping[idx] = _STAR_HEADERS[header]
    return mapping


def _merge_star(star: dict) -> str:
    parts = []
    for key, label in (("situation", "Situation"), ("task", "Task"),
                       ("action", "Action"), ("result", "Result"), ("tips", "Tips")):
        value = str(star.get(key) or "").strip()
        if value:
            parts.append(f"{label}: {value}")
    return "\n".join(parts)


class ExcelStore:
    def __init__(self, path: Path | str, backups_dir: Path | str | None = None):
        self.path = Path(path)
        self.backups_dir = Path(backups_dir) if backups_dir else self.path.parent / "backups"
        self._lock = threading.RLock()
        self._apps: list[dict] = []
        self._prep: list[dict] = []
        self._meta: dict = {}
        self._sig: tuple | None = None
        self._last_backup = 0.0
        if self.path.exists():
            self._load()
        else:
            self.path.parent.mkdir(parents=True, exist_ok=True)
            wb = schema.init_workbook()
            self._meta = {
                "schema_version": schema.SCHEMA_VERSION,
                "next_app_id": 1,
                "next_prep_id": 1,
            }
            self._atomic_save(wb)
            self._load()

    # ---------- file plumbing ----------

    def _signature(self) -> tuple | None:
        try:
            st = self.path.stat()
            return (st.st_mtime_ns, st.st_size)
        except OSError:
            return None

    def _sentinel(self) -> Path:
        return self.path.parent / f"~${self.path.name}"

    def is_locked(self) -> bool:
        return self._sentinel().exists()

    def _load(self) -> None:
        with self._lock:
            wb = load_workbook(self.path, data_only=True)
            apps, prep, meta = [], [], {}
            if schema.SHEET_APPS in wb.sheetnames:
                ws = wb[schema.SHEET_APPS]
                mapping = schema.header_map(ws, schema.APP_COLUMNS)
                for row in ws.iter_rows(min_row=2):
                    rec = {mapping[c.column]: c.value for c in row if c.column in mapping}
                    if rec.get("company") or rec.get("title"):
                        apps.append(_clean(rec, schema.APP_KEYS))
            if schema.SHEET_PREP in wb.sheetnames:
                ws = wb[schema.SHEET_PREP]
                mapping = schema.header_map(ws, schema.PREP_COLUMNS)
                star_map = _legacy_prep_columns(ws)  # pre-v3 STAR-format sheets
                for row in ws.iter_rows(min_row=2):
                    rec = {mapping[c.column]: c.value for c in row if c.column in mapping}
                    if star_map and not rec.get("answer"):
                        star = {star_map[c.column]: c.value for c in row if c.column in star_map}
                        rec["answer"] = _merge_star(star)
                    if rec.get("question"):
                        prep.append(_clean(rec, schema.PREP_KEYS))
            if schema.SHEET_META in wb.sheetnames:
                meta = schema.read_meta(wb[schema.SHEET_META])
            wb.close()

            # Self-heal IDs for rows imported without them.
            next_app = int(meta.get("next_app_id") or 1)
            for rec in apps:
                if not rec.get("id"):
                    rec["id"] = next_app
                    next_app += 1
            next_app = max([next_app] + [int(r["id"]) + 1 for r in apps if r.get("id")])
            next_prep = int(meta.get("next_prep_id") or 1)
            for rec in prep:
                if not rec.get("id"):
                    rec["id"] = next_prep
                    next_prep += 1
            next_prep = max([next_prep] + [int(r["id"]) + 1 for r in prep if r.get("id")])

            meta["next_app_id"] = next_app
            meta["next_prep_id"] = next_prep
            needs_upgrade = int(meta.get("schema_version") or 0) < schema.SCHEMA_VERSION
            meta["schema_version"] = schema.SCHEMA_VERSION
            self._apps, self._prep, self._meta = apps, prep, meta
            self._sig = self._signature()
            if needs_upgrade:
                # Persist the v3 layout (collapsed statuses, merged prep,
                # dropped salary columns). Best-effort: skip if Excel has it open.
                try:
                    self._save()
                except WorkbookLockedError:
                    pass

    def _maybe_reload(self) -> None:
        with self._lock:
            if self._signature() != self._sig:
                self._load()

    def _backup(self) -> None:
        now = time.time()
        if now - self._last_backup < BACKUP_MIN_INTERVAL or not self.path.exists():
            return
        self.backups_dir.mkdir(parents=True, exist_ok=True)
        stamp = datetime.now().strftime("%Y%m%d-%H%M%S")
        shutil.copy2(self.path, self.backups_dir / f"{self.path.stem}-{stamp}.xlsx")
        self._last_backup = now
        backups = sorted(self.backups_dir.glob(f"{self.path.stem}-*.xlsx"))
        for old in backups[:-BACKUP_KEEP]:
            old.unlink(missing_ok=True)

    def _atomic_save(self, wb) -> None:
        tmp = self.path.with_suffix(".xlsx.tmp")
        try:
            wb.save(tmp)
            os.replace(tmp, self.path)
        except PermissionError as exc:
            raise WorkbookLockedError(str(exc)) from exc
        finally:
            if tmp.exists():
                tmp.unlink(missing_ok=True)
            wb.close()

    def _save(self) -> None:
        """Rewrite the whole workbook from memory. Caller must hold the lock."""
        if self.is_locked():
            raise WorkbookLockedError(f"{self.path.name} is open in Excel — close it and retry.")
        self._backup()
        # Rebuild from a fresh template so styling stays canonical.
        wb = schema.init_workbook()
        schema.write_rows(wb[schema.SHEET_APPS], schema.APP_COLUMNS, self._apps)
        schema.write_rows(wb[schema.SHEET_PREP], schema.PREP_COLUMNS, self._prep)
        schema.set_meta(wb[schema.SHEET_META], self._meta)
        self._atomic_save(wb)
        self._sig = self._signature()

    # ---------- applications ----------

    def list_applications(self) -> list[dict]:
        self._maybe_reload()
        with self._lock:
            return [dict(r) for r in self._apps]

    def get_application(self, app_id: int) -> dict:
        self._maybe_reload()
        with self._lock:
            for rec in self._apps:
                if rec["id"] == app_id:
                    return dict(rec)
        raise NotFoundError(f"application {app_id}")

    def add_application(self, fields: dict, *, force_id: int | None = None) -> dict:
        with self._lock:
            self._maybe_reload()
            rec = _clean(fields, schema.APP_KEYS)
            if force_id is not None:
                rec["id"] = force_id
                self._meta["next_app_id"] = max(int(self._meta["next_app_id"]), force_id + 1)
            else:
                rec["id"] = int(self._meta["next_app_id"])
                self._meta["next_app_id"] = rec["id"] + 1
            rec["last_updated"] = datetime.now().isoformat(timespec="seconds")
            self._apps.append(rec)
            self._save()
            return dict(rec)

    def update_application(self, app_id: int, patch: dict) -> tuple[dict, dict]:
        with self._lock:
            self._maybe_reload()
            for i, rec in enumerate(self._apps):
                if rec["id"] == app_id:
                    before = dict(rec)
                    merged = {**rec, **{k: v for k, v in patch.items() if k in schema.APP_KEYS and k != "id"}}
                    merged = _clean(merged, schema.APP_KEYS)
                    merged["id"] = app_id
                    merged["last_updated"] = datetime.now().isoformat(timespec="seconds")
                    self._apps[i] = merged
                    self._save()
                    return before, dict(merged)
        raise NotFoundError(f"application {app_id}")

    def delete_application(self, app_id: int) -> dict:
        with self._lock:
            self._maybe_reload()
            for i, rec in enumerate(self._apps):
                if rec["id"] == app_id:
                    removed = self._apps.pop(i)
                    self._save()
                    return dict(removed)
        raise NotFoundError(f"application {app_id}")

    # ---------- prep ----------

    def list_prep(self) -> list[dict]:
        self._maybe_reload()
        with self._lock:
            return [dict(r) for r in self._prep]

    def add_prep(self, fields: dict, *, force_id: int | None = None) -> dict:
        with self._lock:
            self._maybe_reload()
            rec = _clean(fields, schema.PREP_KEYS)
            if force_id is not None:
                rec["id"] = force_id
                self._meta["next_prep_id"] = max(int(self._meta["next_prep_id"]), force_id + 1)
            else:
                rec["id"] = int(self._meta["next_prep_id"])
                self._meta["next_prep_id"] = rec["id"] + 1
            self._prep.append(rec)
            self._save()
            return dict(rec)

    def update_prep(self, prep_id: int, patch: dict) -> tuple[dict, dict]:
        with self._lock:
            self._maybe_reload()
            for i, rec in enumerate(self._prep):
                if rec["id"] == prep_id:
                    before = dict(rec)
                    merged = {**rec, **{k: v for k, v in patch.items() if k in schema.PREP_KEYS and k != "id"}}
                    merged = _clean(merged, schema.PREP_KEYS)
                    merged["id"] = prep_id
                    self._prep[i] = merged
                    self._save()
                    return before, dict(merged)
        raise NotFoundError(f"prep {prep_id}")

    def delete_prep(self, prep_id: int) -> dict:
        with self._lock:
            self._maybe_reload()
            for i, rec in enumerate(self._prep):
                if rec["id"] == prep_id:
                    removed = self._prep.pop(i)
                    self._save()
                    return dict(removed)
        raise NotFoundError(f"prep {prep_id}")

    # ---------- bulk (geocode backfill) ----------

    def bulk_update_applications(self, patches: dict[int, dict]) -> int:
        """Apply {id: patch} in one save. Returns number of rows changed."""
        with self._lock:
            self._maybe_reload()
            changed = 0
            for i, rec in enumerate(self._apps):
                patch = patches.get(rec["id"])
                if patch:
                    merged = {**rec, **{k: v for k, v in patch.items() if k in schema.APP_KEYS and k != "id"}}
                    self._apps[i] = _clean(merged, schema.APP_KEYS)
                    self._apps[i]["id"] = rec["id"]
                    changed += 1
            if changed:
                self._save()
            return changed

    def bulk_add(self, apps: list[dict], prep: list[dict]) -> tuple[int, int]:
        """Append many records in a single save (importer path)."""
        with self._lock:
            self._maybe_reload()
            next_app = int(self._meta["next_app_id"])
            for fields in apps:
                rec = _clean(fields, schema.APP_KEYS)
                rec["id"] = next_app
                next_app += 1
                if not rec.get("last_updated"):
                    rec["last_updated"] = datetime.now().isoformat(timespec="seconds")
                self._apps.append(rec)
            next_prep = int(self._meta["next_prep_id"])
            for fields in prep:
                rec = _clean(fields, schema.PREP_KEYS)
                rec["id"] = next_prep
                next_prep += 1
                self._prep.append(rec)
            self._meta["next_app_id"] = next_app
            self._meta["next_prep_id"] = next_prep
            if apps or prep:
                self._save()
            return len(apps), len(prep)

    # ---------- info ----------

    def info(self) -> dict:
        self._maybe_reload()
        with self._lock:
            return {
                "xlsx_path": str(self.path),
                "xlsx_exists": self.path.exists(),
                "locked": self.is_locked(),
                "schema_version": self._meta.get("schema_version"),
                "app_count": len(self._apps),
                "prep_count": len(self._prep),
            }
