from datetime import date, datetime

from openpyxl import Workbook

from tracker.excel.legacy import import_legacy_workbook, read_legacy_workbook
from tracker.excel.store import ExcelStore


def build_legacy_file(path):
    wb = Workbook()
    ws = wb.active
    ws.title = "💼 Job Applications"
    ws.append(["MISSION CONTROL — JOB SEARCH OS"])  # title row above headers
    ws.append(["#", "Company", "Job Title", "Date Applied", "Status", "Location",
               "Work Type", "Salary Min (€)", "Salary Max (€)", "Source",
               "Work Permit / Sponsorship", "Referral Contact", "Job Posting URL",
               "Applicant Portal", "Last Updated", "Notes"])
    ws.append([1, "Adyen", "Data Analyst", datetime(2026, 4, 8), "rejected",
               "Amsterdam, Netherlands", "Hybrid", 55000, 70000, "LinkedIn",
               "Recognised sponsor", "", "https://example.com/job/1", "", "", "Note A"])
    ws.append([None, "ASML", "ML Engineer", "2026-05-16T00:00:00.000Z", "Applied",
               "", "", None, None, "", "", "", "https://example.com/job/2", "", "", 85])
    ws.append([None, "", "", None, "", "", "", None, None, "", "", "", "", "", "", ""])  # blank

    prep = wb.create_sheet("❓ Interview Prep")
    prep.append(["PREP BANK"])
    prep.append(["Category", "Sub-Category", "Question", "Situation", "Task",
                 "Action", "Result", "Tips / Key Points"])
    prep.append(["Behavioral", "Teamwork", "Tell me about a conflict", "S", "T", "A", "R", "Be concise"])

    wb.create_sheet("💰 Offer Comparison").append(["Company", "Base Salary (€)"])
    wb.save(path)
    return path


def test_read_legacy(tmp_path):
    src = build_legacy_file(tmp_path / "legacy.xlsx")
    data = read_legacy_workbook(src)
    assert len(data["applications"]) == 2
    first, second = data["applications"]
    assert first["company"] == "Adyen"
    assert first["date_applied"] == "2026-04-08"
    assert first["status"] == "Rejected"          # canonicalized casing
    assert first["geo_status"] == "pending"       # has location, no coords
    assert second["date_applied"] == "2026-05-16"  # SheetJS ISO string
    assert second["notes"] == "85"                 # stray numeric note cast to str
    assert second.get("geo_status", "") == ""      # no location
    assert len(data["prep"]) == 1
    # STAR columns collapse into a single labeled answer
    answer = data["prep"][0]["answer"]
    assert answer.startswith("Situation: S")
    assert answer.endswith("Tips: Be concise")


def test_legacy_stage_names_collapse(tmp_path):
    from openpyxl import Workbook

    path = tmp_path / "stages.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "💼 Job Applications"
    ws.append(["#", "Company", "Job Title", "Date Applied", "Status"])
    ws.append([1, "Acme", "Analyst", "2026-05-01", "Phone Screen"])
    ws.append([2, "Beta", "Engineer", "2026-05-02", "onsite"])
    wb.save(path)

    data = read_legacy_workbook(path)
    assert [r["status"] for r in data["applications"]] == ["Interview", "Interview"]


def test_import_into_store(tmp_path):
    src = build_legacy_file(tmp_path / "legacy.xlsx")
    store = ExcelStore(tmp_path / "tracker.xlsx")
    summary = import_legacy_workbook(src, store)
    assert summary["applications"] == 2
    assert summary["prep"] == 1

    rows = store.list_applications()
    assert [r["id"] for r in rows] == [1, 2]      # fresh sequential IDs
    assert rows[0]["date_applied"] == date(2026, 4, 8)

    # Destination re-loadable from disk.
    fresh = ExcelStore(store.path)
    assert len(fresh.list_applications()) == 2
    assert len(fresh.list_prep()) == 1
