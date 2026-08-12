"""Workbook schema v2: sheet layout, columns, and styling for a fresh tracker file.

The app owns this format (headers on row 1, no emoji sheet names). Legacy
JAMIE-era workbooks are converted once by `waypoint.excel.legacy`.
"""
from __future__ import annotations

from datetime import date, datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

SCHEMA_VERSION = 2

SHEET_APPS = "Applications"
SHEET_PREP = "Interview Prep"
SHEET_META = "Meta"

STATUSES = [
    "Wishlist",
    "Applied",
    "Phone Screen",
    "Technical",
    "Onsite",
    "Offer",
    "Rejected",
    "Withdrawn",
]
ACTIVE_STATUSES = ["Applied", "Phone Screen", "Technical", "Onsite", "Offer"]
WORK_TYPES = ["Onsite", "Hybrid", "Remote"]
GEO_STATUSES = ["ok", "pending", "failed", "remote", "manual"]

# (header, key, width, number_format)
APP_COLUMNS = [
    ("ID", "id", 6, "0"),
    ("Company", "company", 22, None),
    ("Job Title", "title", 32, None),
    ("Status", "status", 13, None),
    ("Date Applied", "date_applied", 13, "yyyy-mm-dd"),
    ("Location", "location", 24, None),
    ("Work Type", "work_type", 10, None),
    ("Salary Min", "salary_min", 11, "#,##0"),
    ("Salary Max", "salary_max", 11, "#,##0"),
    ("Currency", "currency", 9, None),
    ("Source", "source", 14, None),
    ("Sponsorship", "sponsorship", 16, None),
    ("Referral", "referral", 14, None),
    ("Job URL", "url", 40, None),
    ("Portal URL", "portal_url", 28, None),
    ("Notes", "notes", 40, None),
    ("Latitude", "latitude", 10, "0.00000"),
    ("Longitude", "longitude", 10, "0.00000"),
    ("Geo Status", "geo_status", 10, None),
    ("Last Updated", "last_updated", 19, None),
]

PREP_COLUMNS = [
    ("ID", "id", 6, "0"),
    ("Category", "category", 16, None),
    ("Sub-Category", "subcategory", 16, None),
    ("Question", "question", 44, None),
    ("Situation", "situation", 40, None),
    ("Task", "task", 40, None),
    ("Action", "action", 40, None),
    ("Result", "result", 40, None),
    ("Tips", "tips", 40, None),
]

APP_KEYS = [k for _, k, _, _ in APP_COLUMNS]
PREP_KEYS = [k for _, k, _, _ in PREP_COLUMNS]
NUMERIC_KEYS = {"salary_min", "salary_max", "latitude", "longitude"}

_HEADER_FILL = PatternFill("solid", start_color="FF101828")
_HEADER_FONT = Font(name="Calibri", bold=True, color="FFE8EEF9", size=11)


def _style_sheet(ws: Worksheet, columns: list) -> None:
    for idx, (header, _key, width, _fmt) in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=idx, value=header)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(vertical="center")
        ws.column_dimensions[get_column_letter(idx)].width = width
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}1"


def write_rows(ws: Worksheet, columns: list, rows: list) -> None:
    """Replace all data rows (row 2+) with `rows` (list of dicts)."""
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)
    for r, rec in enumerate(rows, start=2):
        for c, (_header, key, _width, fmt) in enumerate(columns, start=1):
            value = rec.get(key)
            if value in ("", None):
                continue
            cell = ws.cell(row=r, column=c, value=value)
            if fmt:
                cell.number_format = fmt
            if isinstance(value, (date, datetime)):
                cell.number_format = "yyyy-mm-dd"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}{max(1, len(rows) + 1)}"


def init_workbook() -> Workbook:
    wb = Workbook()
    ws_apps = wb.active
    ws_apps.title = SHEET_APPS
    _style_sheet(ws_apps, APP_COLUMNS)

    ws_prep = wb.create_sheet(SHEET_PREP)
    _style_sheet(ws_prep, PREP_COLUMNS)

    ws_meta = wb.create_sheet(SHEET_META)
    ws_meta.sheet_state = "hidden"
    set_meta(
        ws_meta,
        {
            "schema_version": SCHEMA_VERSION,
            "next_app_id": 1,
            "next_prep_id": 1,
            "created_at": datetime.now().isoformat(timespec="seconds"),
            "generator": "waypoint",
        },
    )
    return wb


def read_meta(ws: Worksheet) -> dict:
    meta = {}
    for row in ws.iter_rows(min_row=1, max_col=2, values_only=True):
        if row and row[0]:
            meta[str(row[0])] = row[1]
    return meta


def set_meta(ws: Worksheet, meta: dict) -> None:
    if ws.max_row >= 1:
        ws.delete_rows(1, ws.max_row)
    for r, (k, v) in enumerate(sorted(meta.items()), start=1):
        ws.cell(row=r, column=1, value=k)
        ws.cell(row=r, column=2, value=v)


def header_map(ws: Worksheet, columns: list) -> dict:
    """Map column index -> record key based on the actual header row (order-tolerant)."""
    by_header = {h: k for h, k, _, _ in columns}
    mapping = {}
    for idx, cell in enumerate(ws[1], start=1):
        if cell.value and str(cell.value).strip() in by_header:
            mapping[idx] = by_header[str(cell.value).strip()]
    return mapping
