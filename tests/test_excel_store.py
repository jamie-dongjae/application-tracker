from datetime import date, datetime

import pytest
from openpyxl import load_workbook

from tracker.excel.store import ExcelStore, WorkbookLockedError, coerce_date


def test_creates_workbook_on_first_run(tmp_path):
    path = tmp_path / "tracker.xlsx"
    ExcelStore(path)
    wb = load_workbook(path)
    assert set(wb.sheetnames) == {"Applications", "Interview Prep", "Meta"}
    assert wb["Applications"]["A1"].value == "ID"
    assert wb["Meta"].sheet_state == "hidden"


def test_roundtrip_add_update_delete(store):
    rec = store.add_application({"company": "Acme", "title": "Analyst",
                                 "status": "Applied", "date_applied": "2026-08-01"})
    assert rec["id"] == 1
    assert rec["date_applied"] == date(2026, 8, 1)

    before, after = store.update_application(1, {"status": "Interview"})
    assert before["status"] == "Applied" and after["status"] == "Interview"

    # Reload from disk: a fresh store must see the same data.
    fresh = ExcelStore(store.path)
    rows = fresh.list_applications()
    assert len(rows) == 1
    assert rows[0]["status"] == "Interview"

    fresh.delete_application(1)
    assert fresh.list_applications() == []
    # IDs are never reused.
    rec2 = fresh.add_application({"company": "Beta", "title": "Engineer"})
    assert rec2["id"] == 2


def test_v2_workbook_migrates_on_load(tmp_path):
    """Pre-v3 workbooks (old stage names, STAR prep columns) upgrade in place."""
    from openpyxl import Workbook

    path = tmp_path / "tracker.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Applications"
    ws.append(["ID", "Company", "Job Title", "Status", "Date Applied", "Location",
               "Salary Min", "Salary Max", "Currency", "Last Updated"])
    ws.append([1, "Acme", "Analyst", "Phone Screen", "2026-06-01", "Utrecht", 50000, 60000, "EUR", ""])
    ws.append([2, "Beta", "Engineer", "Onsite", "2026-06-02", "", None, None, "", ""])
    prep_ws = wb.create_sheet("Interview Prep")
    prep_ws.append(["ID", "Category", "Question", "Situation", "Task", "Action", "Result", "Tips"])
    prep_ws.append([1, "Behavioral", "Conflict story?", "S", "T", "A", "R", "Breathe"])
    meta_ws = wb.create_sheet("Meta")
    meta_ws.append(["schema_version", 2])
    meta_ws.append(["next_app_id", 3])
    meta_ws.append(["next_prep_id", 2])
    wb.save(path)

    store = ExcelStore(path)
    rows = {r["id"]: r for r in store.list_applications()}
    assert rows[1]["status"] == "Interview"
    assert rows[2]["status"] == "Interview"
    prep = store.list_prep()
    assert prep[0]["answer"] == "Situation: S\nTask: T\nAction: A\nResult: R\nTips: Breathe"

    # The upgrade was persisted: salary columns are gone, version bumped.
    wb2 = load_workbook(path)
    headers = [c.value for c in wb2["Applications"][1]]
    assert "Salary Min" not in headers
    meta = dict(row[:2] for row in wb2["Meta"].iter_rows(values_only=True))
    assert meta["schema_version"] == 3


def test_atomic_save_leaves_no_tmp(store):
    store.add_application({"company": "Acme", "title": "Analyst"})
    assert not list(store.path.parent.glob("*.tmp"))


def test_lock_sentinel_blocks_writes(store):
    sentinel = store.path.parent / f"~${store.path.name}"
    sentinel.write_text("locked by excel")
    with pytest.raises(WorkbookLockedError):
        store.add_application({"company": "Acme", "title": "Analyst"})
    sentinel.unlink()
    rec = store.add_application({"company": "Acme", "title": "Analyst"})
    assert rec["id"] >= 1


def test_backup_created(tmp_path):
    store = ExcelStore(tmp_path / "tracker.xlsx")
    store.add_application({"company": "Acme", "title": "Analyst"})
    backups = list((tmp_path / "backups").glob("tracker-*.xlsx"))
    assert len(backups) == 1  # throttled to at most one per hour


def test_external_edit_is_picked_up(store):
    store.add_application({"company": "Acme", "title": "Analyst"})
    # Simulate the user editing the workbook directly in Excel.
    wb = load_workbook(store.path)
    ws = wb["Applications"]
    headers = [c.value for c in ws[1]]
    row = [""] * len(headers)
    row[headers.index("ID")] = 99
    row[headers.index("Company")] = "HandEdited"
    row[headers.index("Job Title")] = "CFO"
    row[headers.index("Status")] = "Applied"
    ws.append(row)
    wb.save(store.path)

    companies = {r["company"] for r in store.list_applications()}
    assert "HandEdited" in companies


def test_bulk_update(store):
    a = store.add_application({"company": "Acme", "title": "Analyst"})
    b = store.add_application({"company": "Beta", "title": "Engineer"})
    changed = store.bulk_update_applications({
        a["id"]: {"latitude": 52.1, "longitude": 5.1, "geo_status": "ok"},
        b["id"]: {"geo_status": "failed"},
    })
    assert changed == 2
    rows = {r["id"]: r for r in store.list_applications()}
    assert rows[a["id"]]["latitude"] == 52.1
    assert rows[b["id"]]["geo_status"] == "failed"


@pytest.mark.parametrize("value,expected", [
    ("2026-04-08T00:00:00.000Z", date(2026, 4, 8)),   # SheetJS ISO string
    (datetime(2026, 4, 8, 0, 0), date(2026, 4, 8)),
    (45000, date(2023, 3, 15)),                        # Excel serial
    ("08/04/2026", date(2026, 4, 8)),                  # d/m/Y wins for EU data
    ("", None),
    (None, None),
    ("not a date", None),
])
def test_coerce_date(value, expected):
    assert coerce_date(value) == expected
