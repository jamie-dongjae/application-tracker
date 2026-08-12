"""One-time importer for legacy (JAMIE-era) workbooks.

Legacy shape: emoji sheet names ("💼 Job Applications"), a title row above the
headers, free-form column names. The source file is opened read-only and never
modified.
"""
from __future__ import annotations

import re
from pathlib import Path

from openpyxl import load_workbook

from .schema import STATUSES
from .store import ExcelStore, coerce_date

# Ordered (pattern, key) — first match wins. Patterns test the normalized header.
APP_ALIASES = [
    (r"^company$", "company"),
    (r"^(job )?title$|^role$|^position$", "title"),
    (r"^date applied$|^applied", "date_applied"),
    (r"^status$", "status"),
    (r"^location$|^city$", "location"),
    (r"^work ?(type|mode)$|^arrangement$", "work_type"),
    (r"^salary min", "salary_min"),
    (r"^salary max", "salary_max"),
    (r"^source$", "source"),
    (r"sponsorship|work permit|^visa", "sponsorship"),
    (r"^referral", "referral"),
    (r"^job (posting )?url$|^url$|^link$", "url"),
    (r"portal", "portal_url"),
    (r"^last updated$", "last_updated"),
    (r"^notes?$", "notes"),
    (r"^lat(itude)?$", "latitude"),
    (r"^(lng|lon(gitude)?)$", "longitude"),
]

PREP_ALIASES = [
    (r"^category$", "category"),
    (r"^sub.?category$", "subcategory"),
    (r"^question$", "question"),
    (r"^situation$", "situation"),
    (r"^task$", "task"),
    (r"^action$", "action"),
    (r"^result$", "result"),
    (r"^tips", "tips"),
]

_CANON_STATUS = {s.lower(): s for s in STATUSES}


def _norm_header(value) -> str:
    text = str(value or "").strip().lower()
    text = re.sub(r"\(.*?\)", "", text)  # drop units like "(€)" / "(1-5)"
    return re.sub(r"\s+", " ", text).strip()


def _match_key(header: str, aliases: list) -> str | None:
    for pattern, key in aliases:
        if re.search(pattern, header):
            return key
    return None


def _find_sheet(wb, needle: str):
    for name in wb.sheetnames:
        if needle in name.lower():
            return wb[name]
    return None


def _detect_columns(ws, aliases: list, min_hits: int = 3):
    """Scan the first rows for the header row; return (header_row, {col: key})."""
    for row_idx in range(1, min(9, ws.max_row + 1)):
        mapping = {}
        for col_idx, cell in enumerate(ws[row_idx], start=1):
            key = _match_key(_norm_header(cell.value), aliases) if cell.value else None
            if key and key not in mapping.values():
                mapping[col_idx] = key
        if len(mapping) >= min_hits:
            return row_idx, mapping
    return None, {}


def read_legacy_workbook(src: Path) -> dict:
    """Extract applications + prep rows from a legacy workbook (no writes)."""
    wb = load_workbook(src, data_only=True, read_only=True)
    apps, prep, warnings = [], [], []

    ws = _find_sheet(wb, "job applications")
    if ws is None:
        warnings.append("no 'Job Applications' sheet found")
    else:
        header_row, mapping = _detect_columns(ws, APP_ALIASES)
        if not mapping:
            warnings.append(f"could not detect headers in sheet '{ws.title}'")
        else:
            for row in ws.iter_rows(min_row=header_row + 1):
                rec = {mapping[c.column]: c.value for c in row
                       if getattr(c, "column", None) in mapping and c.value is not None}
                if not str(rec.get("company") or "").strip():
                    continue
                applied = coerce_date(rec.get("date_applied"))
                rec["date_applied"] = applied.isoformat() if applied else ""
                status = str(rec.get("status") or "").strip()
                rec["status"] = _CANON_STATUS.get(status.lower(), status or "Applied")
                rec["notes"] = str(rec.get("notes") or "").strip()
                if rec.get("salary_min") or rec.get("salary_max"):
                    rec.setdefault("currency", "EUR")
                if rec.get("latitude") and rec.get("longitude"):
                    rec["geo_status"] = "ok"
                elif str(rec.get("location") or "").strip():
                    rec["geo_status"] = "pending"
                apps.append(rec)

    ws = _find_sheet(wb, "interview prep")
    if ws is not None:
        header_row, mapping = _detect_columns(ws, PREP_ALIASES)
        for row in ws.iter_rows(min_row=(header_row or 0) + 1):
            rec = {mapping[c.column]: c.value for c in row
                   if getattr(c, "column", None) in mapping and c.value is not None}
            if str(rec.get("question") or "").strip():
                prep.append(rec)

    wb.close()
    return {"applications": apps, "prep": prep, "warnings": warnings}


def import_legacy_workbook(src: Path, store: ExcelStore) -> dict:
    data = read_legacy_workbook(src)
    n_apps, n_prep = store.bulk_add(data["applications"], data["prep"])
    return {"applications": n_apps, "prep": n_prep, "warnings": data["warnings"],
            "source": str(src)}
